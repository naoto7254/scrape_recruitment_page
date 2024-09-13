import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook

import scraper as sc

class ExcelProcessor:
    def __init__(self, file_path):
        self.df = pd.read_excel(file_path)
    

    def input_data(self):
        elements_data = self.df.iloc[:1, :]
        company_name = self.df.iloc[2, 1]
        urls = self.df.iloc[5:, 0]
        
        return elements_data, company_name, urls
    
    def processing_data_for_scraping(self):
        elements_data, company_name, urls = self.input_data()
        
        header_array = elements_data.columns.tolist()
        urls_array = urls.tolist()
        elements_info_array = elements_data.values.tolist()
        
        search_elements_array = []
        
        for element in elements_info_array:
            # NaN 以外の値を含む辞書を作成
            element_dict = {key: (value,) for key, value in zip(header_array, element) if not pd.isna(value)}
            search_elements_array.append(element_dict)
            
        del search_elements_array[0]['No,'] # 配列の先頭を削除>['No.': '取得要素▶']が入っているから
        
        # ここをどうするかだけ
        scraper = sc.Scraper(urls_array, search_elements_array)
        scrape_result = scraper.scrape_urls()
        
        return header_array, urls_array, search_elements_array, scrape_result
    
    def processing_data_for_save_excel(self, scrape_result, header_array, urls_array):
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Output"

        header_column_mapping = [(header_array, self.get_column_letters(index)) for index, header_array in enumerate(header_array)]
        header_to_column = dict(header_column_mapping)
        
        elements_header = scrape_result[0]
        elements_values = scrape_result[1:]
        
        ws.append(header_array)
        
        for index, item in enumerate(elements_header):
            for idx, value in enumerate(elements_values):
                ws[f'{header_to_column[item]}{idx + 2}'] = value[index]
        
        for index, url in enumerate(urls_array, start=2):
            ws[f'{header_to_column["HP_URL"]}{index}'] = url
            
        return wb
    
    
    def get_column_letters(self, num): # 指定した番号に対応する列のアルファベットを返す
        letters = ''
        while num >= 0:
            letters = chr(num % 26 + 65) + letters
            num = num // 26 - 1
        return letters
        


    def save_excel(self, output_path, wb):         
        wb.save(output_path)
    
